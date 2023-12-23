# 图像和文档处理程序

from PIL import Image, ImageFilter

# if __name__ == '__main__':
#     image = Image.open('../files/dog01.jpg')
#     image.filter(ImageFilter.CONTOUR).show()

import datetime

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.datetime.now()

wb.save("../files/sample.xlsx")
